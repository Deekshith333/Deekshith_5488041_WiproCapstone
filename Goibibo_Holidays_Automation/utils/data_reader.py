import csv
from pathlib import Path

from config import ROOT_DIR


CSV_PATH = ROOT_DIR / "testdata" / "holiday_test_data.csv"
XLSX_PATH = ROOT_DIR / "testdata" / "holiday_test_data.xlsx"


def read_csv_data(path: Path = CSV_PATH):
    with path.open(newline="", encoding="utf-8") as file:
        return list(csv.DictReader(file))


def read_xlsx_data(path: Path = XLSX_PATH):
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[index]: "" if value is None else str(value) for index, value in enumerate(row)})
    return rows


def get_case(case_id: str, source: str = "csv"):
    rows = read_xlsx_data() if source.lower() == "xlsx" else read_csv_data()
    for row in rows:
        if row["case_id"] == case_id:
            return row
    raise ValueError(f"Test case not found in {source}: {case_id}")


def get_cases(source: str = "csv"):
    return read_xlsx_data() if source.lower() == "xlsx" else read_csv_data()
